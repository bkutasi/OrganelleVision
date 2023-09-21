import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, numbers

class FileHandler:
    def __init__(self, directory):
        self.directory = directory

    def get_files(self):
        return sorted(os.listdir(self.directory))

    def join_path(self, filename):
        return os.path.join(self.directory, filename)

class WorksheetData:
    def __init__(self):
        self.data = {run_names[0]: {"csv_files": [], "image_files": [], "source_image_files": [], "cytosol_csv_files": []},
                     run_names[1]: {"csv_files": [], "image_files": [], "source_image_files": [], "cytosol_csv_files": []}}

    def add_file(self, worksheet_name, file_path, is_image=False, is_source_image=False, is_cytosol=False):
        self.data[worksheet_name]["image_files" if is_image else "source_image_files" if is_source_image else "cytosol_csv_files" if is_cytosol else "csv_files"].append(file_path)

    def get_files(self, worksheet_name, file_type):
        return self.data[worksheet_name][file_type]

class DataHandler:
    @staticmethod
    def get_codes_from_filenames():
        codes = set()
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
        for file_name in os.listdir(path):
            if '-' in file_name:
                code = file_name.split('-')[0] + '-' + file_name.split('-')[1]
                codes.add(code)
        return list(codes)

    def cyto_handler(ws, cytosol_csv_file):
        # Read the cytosol CSV file into a pandas dataframe, and rem. last col
        df_cytosol = pd.read_csv(cytosol_csv_file)

        # Add the base name of the file to the worksheet and enlarge the cell and make the background yellow
        if ws.max_row == 1: cell1 = ws.cell(row=ws.max_row, column=1)
        else: cell1 = ws.cell(row=ws.max_row+1, column=1)
        DataHandler.cyto_cellmerge(ws, cell1, value=os.path.basename(cytosol_csv_file).strip("_cytosol.csv"))

        # Add the name of the cytosol CSV file to the worksheet
        ws.append([os.path.basename(cytosol_csv_file).strip(".csv")])

        # Update previous_cell with the current CSV file name
        previous_cells.append(cytosol_csv_file)

        # Add the cytosol dataframe to the worksheet
        for r in dataframe_to_rows(df_cytosol, index=False, header=True):
            ws.append(r)
            cell1 = ws.cell(row=int(cell1.row), column=1)

        # Add an extra row after the dataframe for spacing
        DataHandler.add_row(ws, cell1)

    def cyto_cellmerge(ws, cell, value):
        # Fills in the value for the naming based on the cytosolic file
        cell.value = value
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(size=14)
        ws.merge_cells(start_row=cell.row, end_row=cell.row, start_column=1, end_column=5)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def add_row(ws):
        # Add an extra row after the dataframe for spacing
        cell = ws.cell(row=ws.max_row, column=1)
        ws.append([""])

    def add_img(ws, img_file, cell):
        img = Image(img_file)
        img.width = img.width * .08
        img.height = img.height * .08
        ws.add_image(img, cell.coordinate)

# Get run and code names (files in the data folder)
run_names=DataHandler.get_codes_from_filenames()

# Directory where the CSV files and TIFF images are located
path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

# Initialize file handler and worksheet data
file_handler = FileHandler(path)
worksheet_data = WorksheetData()

# Sort files into worksheet_data dictionary
for file in file_handler.get_files():
    name = file.split(".")[0]
    if file.endswith("_cytosol.csv"):
        for worksheet_name in [run_names[0], run_names[1]]:
            worksheet_data.add_file(worksheet_name, file_handler.join_path(file), is_cytosol=True)
    elif file.endswith(".ome.tiff.csv"):
        for worksheet_name in [run_names[0], run_names[1]]:
            if file.startswith(worksheet_name):
                worksheet_data.add_file(worksheet_name, file_handler.join_path(file))
    elif file.endswith(".png"):
        for worksheet_name in [run_names[0], run_names[1]]:
            if file.startswith(worksheet_name):
                worksheet_data.add_file(worksheet_name, file_handler.join_path(file), is_image=True)

    elif file.endswith(".jpg"):
        for worksheet_name in [run_names[0], run_names[1]]:
            if file.startswith(worksheet_name):
                worksheet_data.add_file(worksheet_name, file_handler.join_path(file), is_source_image=True)

# Initialize the workbook and worksheets
wb = Workbook()
ws1 = wb.active
ws1.title = run_names[0]
ws2 = wb.create_sheet(title=run_names[1])
ws3 = wb.create_sheet(title="EX-PG_mito1")
ws4 = wb.create_sheet(title="EX-WT_mito1")

# Initialize previous_cells and areas
previous_cells = []
cytosolic_area_sum = 0
mito1_area_sum = 0

# Iterate the worksheet names and worksheet objects
for run_name, ws in [(run_names[0], ws1), (run_names[1], ws2)]:
    # Process the worksheet_data dictionary for each worksheet
    for csv_file, img_file in zip(worksheet_data.get_files(run_name, "csv_files"), worksheet_data.get_files(run_name, "image_files")):
        try:
            # Look for the corresponding cytosol csv file
            for cytosol_csv_file in worksheet_data.get_files(run_name, "cytosol_csv_files"):
                if os.path.basename(csv_file)[:14] == os.path.basename(cytosol_csv_file)[:14] and cytosol_csv_file not in previous_cells:
                    #extra row if not the first for spacing
                    if ws.max_row != 1: ws.append([""])

                    # Read the cytosol CSV file into a pandas dataframe, and rem. last col
                    df_cytosol = pd.read_csv(cytosol_csv_file, index_col=0,header="infer")

                    # Add the base name of the file to the worksheet and enlarge the cell and make the background yellow
                    DataHandler.cyto_cellmerge(ws,
                     cell=ws.cell(row=ws.max_row, column=1),
                     value=os.path.basename(cytosol_csv_file).strip("_cytosol.csv"))

                    # Add the name of the cytosol CSV file to the worksheet
                    ws.append([os.path.basename(cytosol_csv_file).strip(".csv")])

                    # Update previous_cell with the current CSV file name
                    previous_cells.append(cytosol_csv_file)

                    # Add the cytosol dataframe to the worksheet
                    for r in dataframe_to_rows(df_cytosol, index=True, header=True):
                        ws.append(r)
                        
                    cytosolic_area = df_cytosol.iloc[0, 1]
                    cytosolic_area_sum+=cytosolic_area
                    # Add an extra row after the dataframe for spacing
                    DataHandler.add_row(ws)
                    break

            # Add the name of the CSV file to the worksheet
            ws.append([os.path.basename(csv_file).strip(".ome.tiff.csv")])

            # Define annotation image cell
            cell1 = ws.cell(row=ws.max_row, column=10)  # J is 10th column

            # Define source image cell
            cell_source = ws.cell(row=ws.max_row, column=15)  # o is 15th column

            # Add the annotation image to the worksheet
            DataHandler.add_img(ws, img_file, cell1)

            # Calculate the number of rows occupied by the image
            rows_occupied = int(Image(img_file).height *.08 // 20) # Empirically figured out number, lower division=more rows added

            # Add the source image next to the annotation image
            for source_image_file in worksheet_data.get_files(run_name, "source_image_files"):
                if os.path.basename(source_image_file)[:14] == os.path.basename(img_file)[:14]:
                    DataHandler.add_img(ws, source_image_file, cell_source)

            # Read the CSV file into a pandas dataframe
            df = pd.read_csv(csv_file)

            # Add the dataframe to the worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            DataHandler.add_row(ws)
            
            # Compute required values for each column
            sums = df.iloc[:, 1:].sum()
            averages = df.iloc[:, 1:].mean()
            stdevs = df.iloc[:, 1:].std()
            ratio = [sums[0] / cytosolic_area]

            # Get out mito1 are data
            if "mito1" in csv_file: mito1_area_sum += sums[0]

            # Write the values to the worksheet
            row_num = ws.max_row + 1
            headers = ["SUM", "AVG", "STDEV", "MITO - CYTOSOL RATIO"]
            values = [sums, averages, stdevs, ratio]

            for i, header in enumerate(headers):
                ws.cell(row=row_num+i, column=1, value=header)
                for j, val in enumerate(values[i]):
                    ws.cell(row=row_num+i, column=j+2).value = val
                    if header == "MITO - CYTOSOL RATIO":
                        ws.cell(row=row_num+i, column=j+2).number_format = numbers.FORMAT_PERCENTAGE_00

            # Add empty rows if the csv_file is mito csv and number of records are less than the rows occupied
            if csv_file in worksheet_data.get_files(run_name, "csv_files") and len(df.index)+4 < rows_occupied:
                for i in range(rows_occupied - len(df.index)):
                    ws.append([""])
                    
            # Add an extra row after the dataframe for spacing
            DataHandler.add_row(ws)

        except Exception as e:
            print(f"Error in descriptive data processing {csv_file}: {e}")

    # Add header rows to the worksheets
    ws.column_dimensions['A'].width = 27

    # Deleting redundant feret coordinate columns
    ws.delete_cols(6, 3)
    
    # Adding the total area
    ws.cell(1, 6).value="Total cytosol area:"
    ws.cell(1, 8).value=cytosolic_area_sum 
    
    # Adding the total mito1 area
    ws.cell(1, 10).value="Total area occupied by mito1:"
    ws.cell(1, 13).value=mito1_area_sum

    # Adding the ratio
    ws.cell(1, 15).value="Total area ratio - mito1/cyotsol:"
    ws.cell(1, 18).value=mito1_area_sum/cytosolic_area_sum
    ws.cell(1, 18).number_format = numbers.FORMAT_PERCENTAGE_00


# Process the worksheet_data dictionary for creating pure mito1 worksheets
for run_name, ws in [(run_names[0], ws3), (run_names[1], ws4)]:
    for csv_file in worksheet_data.get_files(run_name, "csv_files"):
        try:
            if "mito1" in csv_file:
                # Read the CSV files into a pandas dataframe
                df = pd.read_csv(csv_file)

                # Remove the first column
                df = df.iloc[:, 1:]

                # Define cell
                cell1 = ws.cell(row=ws.max_row, column=1)

                # Add the dataframe to the worksheet
                for r in dataframe_to_rows(df, index=False, header=False):
                    row_num = int(cell1.row)
                    ws.append(r)
                    cell1 = ws.cell(row=row_num, column=1)

        except Exception as e:
            print(f"Error in mito1 processing {csv_file}: {e}")

    # Delete redundant first row
    ws.delete_rows(1)
    
    # Deleting redundant feret coordinate columns
    ws.delete_cols(6, 3)

    # sum Area col
    sum_value = sum(ws.cell(row=i, column=1).value for i in range(1, ws.max_row+1))

    # Insert row for headers
    ws.insert_rows(1)

    # Set the header values for the new row
    headers = ['Area', 'Perim.', 'Circ.', 'Feret', 'MinFeret', 'AR', 'Round', 'Solidity']
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1).value = header

    # insert a new row at the beginning of the worksheet
    ws.insert_rows(1)

    # write the sum value to cell A1
    ws.cell(row=1, column=1, value=sum_value)

    # insert a new row at the beginning of the worksheet
    ws.insert_rows(1)

    # Name the totel cytosol area
    ws.cell(row=1, column=1).value = "Total area occupied by mito1"

# Save the workbook
wb.save("sed_output.xlsx")
